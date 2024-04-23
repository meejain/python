from datetime import date
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from datetime import date
import os
import shutil

src = '/Users/meejain/Downloads/report.csv'
dst_folder = '/Users/meejain/OneDrive/Adobe/Tim Bonnett - BPA Reports/'

customername = input('What is the Customer name? ')
customer=customername.replace(' ', '')
org_id = input('What is the Org ID? ')
dst = customer + '_' + org_id + '_BPA.csv'
os.rename(src, dst_folder + dst)

exl = str(Path.home()) + "/OneDrive/OneDrive - Adobe/BPAReports/basic2cs-tracking.xlsx"
print (exl)
book = openpyxl.load_workbook(exl)
sheet = book['Customer Profile']
today = date.today()

max_row=sheet.max_row
max_column=sheet.max_column
break_out_flag = False

for i in range(1,max_row+1):
    for j in range(1, max_column + 1):
        cell_obj = sheet.cell(row=i, column=j)
        if cell_obj.value == customername:
            checkbox=sheet.cell(row=i, column=6)
            checkbox.value = '☑'
            todaydate=sheet.cell(row=i, column=7)
            todaydate.value = today
            print (customername + " is successfully updated into the sheet")
            break_out_flag = True
            break
        else:
            if i == max_row and j == max_column:
                print (customername + " is not available in the whole sheet. Please update the sheet manually")
            else:
                continue
    if break_out_flag:
        break

book.save(exl)






